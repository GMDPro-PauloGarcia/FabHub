from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ── COLOUR PALETTE ────────────────────────────────────────────────────────────
NAVY    = "1E293B"
AMBER   = "F59E0B"
AMBER_L = "FEF3C7"
GREEN   = "059669"
GREEN_L = "D1FAE5"
BLUE_L  = "EFF6FF"
GRAY_H  = "F1F5F9"
GRAY_B  = "94A3B8"
RED_L   = "FEF2F2"
WHITE   = "FFFFFF"
PURPLE_L= "F5F3FF"

def fill(hex_):  return PatternFill("solid", fgColor=hex_)
def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)
def thick_bottom():
    return Border(bottom=Side(style="medium"))

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – SALES IMPORT
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "📋 Sales Import"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A4"   # freeze rows 1-3 (title + header)

# ── Title row ─────────────────────────────────────────────────────────────────
ws.merge_cells("A1:T1")
ws["A1"] = "GMD PRO · FABHUB SMART IMPORT  —  Sales & Projects"
ws["A1"].fill      = fill(NAVY)
ws["A1"].font      = Font(bold=True, color=AMBER, size=14, name="Calibri")
ws["A1"].alignment = center()
ws.row_dimensions[1].height = 34

# ── Sub-title / instructions row ─────────────────────────────────────────────
ws.merge_cells("A2:T2")
ws["A2"] = (
    "Fill in from Row 4 onwards.  "
    "Required fields are marked *.  "
    "For AWARDED projects also fill the green columns (N–T).  "
    "Delete this row and Row 3 before importing."
)
ws["A2"].fill      = fill(AMBER_L)
ws["A2"].font      = Font(color="92400E", size=9, italic=True, name="Calibri")
ws["A2"].alignment = center()
ws.row_dimensions[2].height = 22

# ── Column definitions ────────────────────────────────────────────────────────
# (col_letter, header, width, section, required, example, note)
COLS = [
    # ── DEAL INFO ─────────────────────────────────────────
    ("A", "CE No *",            14, "DEAL",     True,  "CE-2026-001",       "Unique cost-estimate number"),
    ("B", "CE Type *",          20, "DEAL",     True,  "Fabrication / General", "See Valid Values sheet"),
    ("C", "Client Name *",      24, "DEAL",     True,  "SM Retail Inc.",    "Company or individual name"),
    ("D", "Project / Product *",28, "DEAL",     True,  "Custom Shelving – SM Megamall", "Brief project description"),
    ("E", "Stage *",            28, "DEAL",     True,  "05 · Client Approval / Revision", "See Valid Values sheet"),
    ("F", "Contract Value",     16, "DEAL",     False, "850000",            "In PHP, numbers only"),
    ("G", "Sales Owner *",      18, "DEAL",     True,  "Mar",               "First name of AE who owns this"),
    ("H", "Assigned AE",        18, "DEAL",     False, "Mar",               "Account Executive"),
    ("I", "BizDev Source",      20, "DEAL",     False, "Referral – Jun",    "How was this lead found?"),
    ("J", "Date Acquired *",    14, "DEAL",     True,  "2026-01-15",        "YYYY-MM-DD format"),
    ("K", "Due Date",           14, "DEAL",     False, "2026-03-01",        "Proposal / decision deadline"),
    ("L", "Contact Person",     20, "DEAL",     False, "Ana Reyes",         "Client contact name"),
    ("M", "Notes",              30, "DEAL",     False, "Scope TBD pending lease",  "Any relevant notes"),
    # ── AWARDED ONLY ──────────────────────────────────────
    ("N", "Award Trigger",      22, "AWARDED",  False, "Contract Signing",  "Contract Signing / PO Issuance / LOI"),
    ("O", "Trigger Date",       14, "AWARDED",  False, "2026-02-10",        "Date contract/PO was signed"),
    ("P", "Start Date",         14, "AWARDED",  False, "2026-02-17",        "Actual project start date"),
    ("Q", "JO No",              14, "AWARDED",  False, "JO-2026-001",       "Job Order number"),
    ("R", "AE Assigned (JO)",   18, "AWARDED",  False, "Mar",               "AE on the JO"),
    ("S", "Comms Group",        22, "AWARDED",  False, "GM SMM Project GC", "Viber / Telegram group name"),
    ("T", "Payment Status",     16, "AWARDED",  False, "Unpaid",            "Unpaid / Partial / Paid"),
]

SECTION_COLORS = {
    "DEAL":    (NAVY,    WHITE),
    "AWARDED": (GREEN,   WHITE),
}
SECTION_BG = {
    "DEAL":    BLUE_L,
    "AWARDED": GREEN_L,
}

# ── Section header row (row 3) ────────────────────────────────────────────────
ws.row_dimensions[3].height = 28

# Find section spans
sections = {}
for col in COLS:
    sec = col[3]
    if sec not in sections:
        sections[sec] = {"start": col[0], "end": col[0]}
    else:
        sections[sec]["end"] = col[0]

for sec, span in sections.items():
    bg, fg = SECTION_COLORS[sec]
    label = "DEAL INFORMATION" if sec == "DEAL" else "AWARDED PROJECTS ONLY (fill if stage ≥ 06)"
    start_col = span["start"]
    end_col   = span["end"]
    ws.merge_cells(f"{start_col}3:{end_col}3")
    cell = ws[f"{start_col}3"]
    cell.value     = label
    cell.fill      = fill(bg)
    cell.font      = Font(bold=True, color=fg, size=9, name="Calibri")
    cell.alignment = center()

# ── Column headers (row 3 is section, row 4 is headers) ──────────────────────
# Wait — rows are: 1=title, 2=instructions, 3=section headers, 4=col headers
# Let me shift: freeze from row 5
ws.freeze_panes = "A5"

# Re-do: section header on row 3, col header on row 4
ws.row_dimensions[4].height = 42

for i, (col_l, header, width, section, required, example, note) in enumerate(COLS):
    bg, fg = SECTION_COLORS[section]
    cell = ws[f"{col_l}4"]
    cell.value     = header
    cell.fill      = fill(bg)
    cell.font      = Font(bold=True, color=fg, size=9, name="Calibri")
    cell.alignment = center()
    cell.border    = border()
    ws.column_dimensions[col_l].width = width

# ── Example rows (5 & 6) ─────────────────────────────────────────────────────
sample_rows = [
    # Pipeline deal
    [
        "CE-2026-001", "Fabrication / General", "SM Retail Inc.",
        "Custom Shelving – SM North EDSA", "04 · Design & CE in Progress",
        "750000", "Mar", "Mar", "Referral – Jun",
        "2026-01-10", "2026-02-28", "Ana Reyes",
        "Scope includes gondola shelving x 40 units",
        "", "", "", "", "", "", "Unpaid",
    ],
    # Awarded deal
    [
        "CE-2026-002", "Fabrication / General", "Robinsons Malls",
        "Kiosk Installation – Galleria", "08 · Fabrication / Construction",
        "1200000", "Paulo", "Paulo", "Direct — Paulo",
        "2026-01-05", "", "Ben Cruz",
        "Fast-track, target completion April",
        "Contract Signing", "2026-01-20", "2026-01-27",
        "JO-2026-001", "Mar", "GMD Galleria GC", "Partial",
    ],
    # Another pipeline
    [
        "CE-2026-003", "Interior Design", "Ayala Malls",
        "F&B Fit-Out – TriNoma", "02 · Client Engagement",
        "500000", "Anna", "Anna", "Cold Call",
        "2026-01-18", "2026-03-15", "Tina Lim",
        "Awaiting brief from client",
        "", "", "", "", "", "", "",
    ],
]

for r_idx, row_data in enumerate(sample_rows, start=5):
    bg_row = GRAY_H if r_idx % 2 == 1 else WHITE
    for c_idx, (col_l, _, _, section, _, _, _) in enumerate(COLS):
        cell = ws[f"{col_l}{r_idx}"]
        cell.value     = row_data[c_idx] if row_data[c_idx] != "" else None
        cell.font      = font(size=9)
        cell.alignment = left()
        cell.border    = border("thin")
        # Awarded columns get green tint
        if section == "AWARDED":
            cell.fill = fill(GREEN_L if row_data[c_idx] else "F0FDF4")
        else:
            cell.fill = fill(bg_row)

# ── Empty data rows 8–60 ─────────────────────────────────────────────────────
for r in range(8, 61):
    bg_row = GRAY_H if r % 2 == 0 else WHITE
    for col_l, _, _, section, _, _, _ in COLS:
        cell = ws[f"{col_l}{r}"]
        cell.font      = font(size=9)
        cell.alignment = left()
        cell.border    = border("thin")
        cell.fill      = fill(GREEN_L if section == "AWARDED" else bg_row)

# ── Data validation ───────────────────────────────────────────────────────────
stages = (
    '"01 · BizDev,02 · Client Engagement,03 · Design Request & Folder Setup,'
    '04 · Design & CE in Progress,05 · Client Approval / Revision,'
    '06 · Project Kickoff,07 · Budget & Briefing,08 · Fabrication / Construction,'
    '09 · Site Visit & Progress Billing,10 · Installation,11 · Punchlist,'
    '12 · Project Close-Out,13 · Client Feedback,Cancelled"'
)
dv_stage = DataValidation(type="list", formula1=stages, allow_blank=True)
dv_stage.sqref = "E5:E200"
ws.add_data_validation(dv_stage)

ce_types = '"Fabrication / General,Interior Design,Renovation,Signage,Fit-Out,Other"'
dv_cetype = DataValidation(type="list", formula1=ce_types, allow_blank=True)
dv_cetype.sqref = "B5:B200"
ws.add_data_validation(dv_cetype)

pay_stat = '"Unpaid,Partial,Paid,Refunded"'
dv_pay = DataValidation(type="list", formula1=pay_stat, allow_blank=True)
dv_pay.sqref = "T5:T200"
ws.add_data_validation(dv_pay)

triggers = '"Contract Signing,PO Issuance,LOI,Verbal Award"'
dv_trig = DataValidation(type="list", formula1=triggers, allow_blank=True)
dv_trig.sqref = "N5:N200"
ws.add_data_validation(dv_trig)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – INSTRUCTIONS
# ══════════════════════════════════════════════════════════════════════════════
wi = wb.create_sheet("📖 Instructions")
wi.sheet_view.showGridLines = False
wi.column_dimensions["A"].width = 22
wi.column_dimensions["B"].width = 18
wi.column_dimensions["C"].width = 40
wi.column_dimensions["D"].width = 36
wi.column_dimensions["E"].width = 14

# Title
wi.merge_cells("A1:E1")
wi["A1"] = "FABHUB SMART IMPORT  ·  Field Guide"
wi["A1"].fill      = fill(NAVY)
wi["A1"].font      = Font(bold=True, color=AMBER, size=13, name="Calibri")
wi["A1"].alignment = center()
wi.row_dimensions[1].height = 32

# Header row
headers2 = ["Column", "Field Name", "Description", "Example / Valid Values", "Required?"]
header_bg = [NAVY, NAVY, NAVY, NAVY, NAVY]
for c, h in enumerate(headers2, 1):
    cell = wi.cell(row=2, column=c)
    cell.value     = h
    cell.fill      = fill(NAVY)
    cell.font      = Font(bold=True, color=WHITE, size=9, name="Calibri")
    cell.alignment = center()
    cell.border    = border()
wi.row_dimensions[2].height = 26

# Field rows
instructions = [
    ("A", "CE No",           "Your internal cost-estimate reference number.",
     "CE-2026-001  /  CE-2026-045", "YES"),
    ("B", "CE Type",         "Category of the project.",
     "Fabrication / General  |  Interior Design  |  Renovation  |  Signage  |  Fit-Out  |  Other", "YES"),
    ("C", "Client Name",     "Company or person name of the client.",
     "SM Retail Inc.  /  Robinsons Malls", "YES"),
    ("D", "Project / Product","Short description of what is being built.",
     "Custom Shelving – SM North EDSA", "YES"),
    ("E", "Stage",           "Current pipeline stage. Use the exact stage name OR just the number (e.g. '05').",
     "01 · BizDev … 13 · Client Feedback  |  Cancelled", "YES"),
    ("F", "Contract Value",  "Total contract amount in PHP. Numbers only, no commas or peso sign.",
     "850000", "No"),
    ("G", "Sales Owner",     "First name of the AE responsible for this deal.",
     "Mar  /  Paulo  /  Anna", "YES"),
    ("H", "Assigned AE",     "Account Executive name (usually same as Sales Owner).",
     "Mar", "No"),
    ("I", "BizDev Source",   "How was this lead generated?",
     "Referral – Jun  /  Cold Call  /  Walk-in  /  Direct – Paulo", "No"),
    ("J", "Date Acquired",   "Date the lead was first entered. Format: YYYY-MM-DD",
     "2026-01-15", "YES"),
    ("K", "Due Date",        "Deadline for proposal or decision. Format: YYYY-MM-DD",
     "2026-03-01", "No"),
    ("L", "Contact Person",  "Name of client point-of-contact.",
     "Ana Reyes", "No"),
    ("M", "Notes",           "Any free-text notes about the deal.",
     "Scope TBD pending lease approval", "No"),
    ("N", "Award Trigger",   "AWARDED ONLY. What document triggered the project award?",
     "Contract Signing  |  PO Issuance  |  LOI  |  Verbal Award", "No"),
    ("O", "Trigger Date",    "AWARDED ONLY. Date the award trigger was signed/issued. YYYY-MM-DD",
     "2026-02-10", "No"),
    ("P", "Start Date",      "AWARDED ONLY. Actual on-site or fabrication start date. YYYY-MM-DD",
     "2026-02-17", "No"),
    ("Q", "JO No",           "AWARDED ONLY. Job Order number assigned to this project.",
     "JO-2026-001", "No"),
    ("R", "AE Assigned (JO)","AWARDED ONLY. AE assigned on the Job Order.",
     "Mar", "No"),
    ("S", "Comms Group",     "AWARDED ONLY. Name of Viber/Telegram group chat for this project.",
     "GMD Galleria GC", "No"),
    ("T", "Payment Status",  "Current payment status of the project.",
     "Unpaid  |  Partial  |  Paid", "No"),
]

for r, (col, name, desc, example, req) in enumerate(instructions, start=3):
    bg = GREEN_L if col in list("NOPQRST") else (BLUE_L if r % 2 == 0 else WHITE)
    req_bg = GREEN_L if req == "YES" else RED_L
    req_color = GREEN if req == "YES" else "DC2626"

    data = [col, name, desc, example, req]
    for c, val in enumerate(data, 1):
        cell = wi.cell(row=r, column=c)
        cell.value     = val
        cell.font      = font(bold=(c == 2), size=9, color=("1E293B" if c != 5 else req_color))
        cell.alignment = left() if c != 1 else center()
        cell.border    = border("thin")
        cell.fill      = fill(req_bg if c == 5 else bg)
    wi.row_dimensions[r].height = 28

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – VALID VALUES
# ══════════════════════════════════════════════════════════════════════════════
vv = wb.create_sheet("✅ Valid Values")
vv.sheet_view.showGridLines = False

vv.merge_cells("A1:D1")
vv["A1"] = "FABHUB SMART IMPORT  ·  Valid Values Reference"
vv["A1"].fill      = fill(NAVY)
vv["A1"].font      = Font(bold=True, color=AMBER, size=13, name="Calibri")
vv["A1"].alignment = center()
vv.row_dimensions[1].height = 32

sections_vv = [
    ("A", "STAGE (Column E)", NAVY, [
        ("01 · BizDev",                        "Lead identified, no contact yet"),
        ("02 · Client Engagement",             "First meeting / calls ongoing"),
        ("03 · Design Request & Folder Setup", "DRF submitted, folder created"),
        ("04 · Design & CE in Progress",       "CE being prepared"),
        ("05 · Client Approval / Revision",    "CE sent, waiting for sign-off"),
        ("06 · Project Kickoff",               "AWARDED — project starts"),
        ("07 · Budget & Briefing",             "Internal budget finalized"),
        ("08 · Fabrication / Construction",    "On the floor / on-site"),
        ("09 · Site Visit & Progress Billing", "Billing milestone reached"),
        ("10 · Installation",                  "Items being installed"),
        ("11 · Punchlist",                     "Defects / snag list"),
        ("12 · Project Close-Out",             "Turnover / COC signed"),
        ("13 · Client Feedback",               "Post-project survey"),
        ("Cancelled",                          "Project cancelled"),
    ]),
    ("C", "CE TYPE (Column B)", GREEN, [
        ("Fabrication / General", "Custom furniture & millwork"),
        ("Interior Design",       "Design-only or design-led project"),
        ("Renovation",            "Fit-out of existing space"),
        ("Signage",               "Signage and branding installation"),
        ("Fit-Out",               "Full retail / commercial fit-out"),
        ("Other",                 "Anything else"),
    ]),
    ("C", "AWARD TRIGGER (Column N)", "7C3AED", [
        ("Contract Signing", "Signed contract document"),
        ("PO Issuance",      "Purchase Order from client"),
        ("LOI",              "Letter of Intent"),
        ("Verbal Award",     "Verbal go-signal (follow up with PO)"),
    ]),
    ("C", "PAYMENT STATUS (Column T)", "0369A1", [
        ("Unpaid",   "No payment received yet"),
        ("Partial",  "Down payment or partial billing paid"),
        ("Paid",     "Fully paid"),
        ("Refunded", "Payment returned to client"),
    ]),
]

vv.column_dimensions["A"].width = 36
vv.column_dimensions["B"].width = 34
vv.column_dimensions["C"].width = 26
vv.column_dimensions["D"].width = 34

row = 3
for start_col, section_title, color, values in sections_vv:
    # Section header
    end_col = "B" if start_col == "A" else "D"
    vv.merge_cells(f"{start_col}{row}:{end_col}{row}")
    cell = vv[f"{start_col}{row}"]
    cell.value     = section_title
    cell.fill      = fill(color)
    cell.font      = Font(bold=True, color=WHITE, size=9, name="Calibri")
    cell.alignment = center()
    vv.row_dimensions[row].height = 24
    row += 1

    for val, desc in values:
        c1 = vv[f"{start_col}{row}"]
        c2_col = "B" if start_col == "A" else "D"
        c2 = vv[f"{c2_col}{row}"]
        c1.value     = val
        c2.value     = desc
        alt = row % 2 == 0
        c1.fill = fill(GRAY_H if alt else WHITE)
        c2.fill = fill(GRAY_H if alt else WHITE)
        c1.font = font(bold=True, size=9)
        c2.font = font(size=9, italic=True, color="475569")
        c1.alignment = left()
        c2.alignment = left()
        c1.border = border("thin")
        c2.border = border("thin")
        vv.row_dimensions[row].height = 22
        row += 1

    row += 1  # blank row between sections

# ── Save ──────────────────────────────────────────────────────────────────────
out = "/home/user/FabHub/GMD_FabHub_SmartImport_Template.xlsx"
wb.save(out)
print(f"Saved: {out}")
